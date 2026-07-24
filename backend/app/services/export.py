import io
from typing import List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from app.models.complaint import Complaint


def generate_excel_export(complaints: List[Complaint]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Complaints Export"

    headers = [
        "Ticket Number", "Title", "Status", "Priority", "Category", 
        "Created At", "Resolved At", "Lot Number", "Product ID", "Customer ID"
    ]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    for c in complaints:
        ws.append([
            c.ticket_number,
            c.title,
            c.status.value,
            c.priority.value,
            c.category.value,
            c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "",
            c.resolved_at.strftime("%Y-%m-%d %H:%M") if c.resolved_at else "",
            c.lot_number or "",
            str(c.product_id) if c.product_id else "",
            str(c.customer_id) if c.customer_id else ""
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def generate_pdf_export(complaint: Complaint) -> io.BytesIO:
    stream = io.BytesIO()
    doc = SimpleDocTemplate(stream, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph(f"Complaint Report: {complaint.ticket_number}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    data = [
        ["Title", complaint.title],
        ["Status", complaint.status.value.upper()],
        ["Priority", complaint.priority.value.upper()],
        ["Category", complaint.category.value.upper()],
        ["Created At", complaint.created_at.strftime("%Y-%m-%d %H:%M:%S") if complaint.created_at else "N/A"],
        ["Resolved At", complaint.resolved_at.strftime("%Y-%m-%d %H:%M:%S") if complaint.resolved_at else "N/A"],
        ["Lot Number", complaint.lot_number or "N/A"],
    ]

    table = Table(data, colWidths=[120, 350])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    desc_title = Paragraph("Description:", styles['Heading3'])
    elements.append(desc_title)
    desc = Paragraph(complaint.description, styles['Normal'])
    elements.append(desc)
    
    if complaint.ai_analysis_data:
        elements.append(Spacer(1, 20))
        ai_title = Paragraph("AI Analysis Data:", styles['Heading3'])
        elements.append(ai_title)
        import json
        ai_text = Paragraph(json.dumps(complaint.ai_analysis_data, indent=2), styles['Code'])
        elements.append(ai_text)

    doc.build(elements)
    stream.seek(0)
    return stream
